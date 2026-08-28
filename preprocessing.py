"""Extract the case documents into a small, citation-friendly JSONL index."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook
from pypdf import PdfReader


CONTROL_PREFIXES = ("00 -", "11 -")
SUPPORTED_SUFFIXES = {".pdf", ".xlsx"}


def make_record(
    source_file: str,
    source_type: str,
    location: dict[str, Any],
    text: str,
) -> dict[str, Any]:
    return {
        "source_file": source_file,
        "source_type": source_type,
        "location": location,
        "text": text.strip(),
    }


def extract_pdf(path: Path) -> Iterator[dict[str, Any]]:
    try:
        reader = PdfReader(path)
    except Exception as exc:
        print(f"Warning: could not read {path.name}: {exc}", file=sys.stderr)
        return

    for page_number, page in enumerate(reader.pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception as exc:
            print(
                f"Warning: could not extract {path.name}, page {page_number}: {exc}",
                file=sys.stderr,
            )
            text = ""

        if not text.strip():
            print(f"Warning: no text extracted from {path.name}, page {page_number}", file=sys.stderr)

        yield make_record(
            path.name,
            "pdf",
            {"page": page_number},
            text,
        )


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value).strip()


def extract_workbook(path: Path) -> Iterator[dict[str, Any]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        print(f"Warning: could not read {path.name}: {exc}", file=sys.stderr)
        return

    try:
        for worksheet in workbook.worksheets:
            rows: list[str] = []
            first_row = None
            last_row = None

            for row in worksheet.iter_rows():
                values = [cell_text(cell.value) for cell in row]
                non_empty = [
                    f"{cell.column_letter}: {value}"
                    for cell, value in zip(row, values)
                    if value
                ]
                if not non_empty:
                    continue

                row_number = row[0].row
                first_row = row_number if first_row is None else first_row
                last_row = row_number
                rows.append(f"Row {row_number}: " + " | ".join(non_empty))

            if not rows:
                print(f"Warning: no values found in {path.name}, sheet {worksheet.title}", file=sys.stderr)
                continue

            yield make_record(
                path.name,
                "xlsx",
                {
                    "sheet": worksheet.title,
                    "row_start": first_row,
                    "row_end": last_row,
                },
                f"Worksheet: {worksheet.title}\n" + "\n".join(rows),
            )
    finally:
        workbook.close()


def source_files(data_dir: Path) -> Iterator[Path]:
    for path in sorted(data_dir.iterdir()):
        if not path.is_file() or path.suffix.lower() not in SUPPORTED_SUFFIXES:
            continue
        if path.name.startswith(CONTROL_PREFIXES):
            continue
        yield path


def extract_documents(data_dir: Path) -> Iterator[dict[str, Any]]:
    for path in source_files(data_dir):
        if path.suffix.lower() == ".pdf":
            yield from extract_pdf(path)
        elif path.suffix.lower() == ".xlsx":
            yield from extract_workbook(path)


def write_index(records: Iterator[dict[str, Any]], output_path: Path) -> int:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    with output_path.open("w", encoding="utf-8") as output_file:
        for record in records:
            json.dump(record, output_file, ensure_ascii=False)
            output_file.write("\n")
            count += 1
    return count


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--data-dir", type=Path, default=Path("Data"))
    parser.add_argument("--output", type=Path, default=Path("working/documents.jsonl"))
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.data_dir.is_dir():
        raise SystemExit(f"Data directory does not exist: {args.data_dir}")

    count = write_index(extract_documents(args.data_dir), args.output)
    print(f"Wrote {count} records to {args.output}")


if __name__ == "__main__":
    main()
